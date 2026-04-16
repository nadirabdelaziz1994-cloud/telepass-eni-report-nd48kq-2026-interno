import os, re, json, sys, shutil, datetime
from pathlib import Path
from collections import defaultdict

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(BASE_DIR, "config_dashboard.json")
TEMPLATE_PATH = os.path.join(BASE_DIR, "template_dashboard.html")
FILE_UTILI_DIR = os.path.join(BASE_DIR, "FILE_UTILI")
GARE_INPUT_DIR = os.path.join(BASE_DIR, "input", "gare")

BLUE = "123764"
LINE = "D8E1EE"
LIGHT = "EEF4FB"

def norm_pdv(v):
    if v is None:
        return None
    s=str(v).strip()
    m=re.findall(r"\d+", s)
    return m[0].zfill(5) if m else None


def safe_num(v):
    if v in (None, "", "-"):
        return None
    try:
        fv=float(v)
        # normalize weird percentages stored as fractions where counts expected
        # keep if near integer or larger than 1
        return fv
    except Exception:
        return None



def load_live_gara():
    folder = Path(GARE_INPUT_DIR)
    if not folder.exists():
        return None
    files = sorted([p for p in folder.rglob("*.xlsx") if not p.name.startswith("~$")], key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        return None
    path = files[0]
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    raw_date = ws.cell(1,4).value
    updated_at = raw_date.strftime("%d/%m/%Y") if hasattr(raw_date, "strftime") else (str(raw_date) if raw_date else "")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        pdv = norm_pdv(row[0] if len(row) > 0 else None)
        if not pdv:
            continue
        vendite = safe_num(row[3] if len(row) > 3 else None)
        rows.append({
            "pdv": pdv,
            "rzv": (row[1] if len(row) > 1 else "") or "",
            "city": (row[2] if len(row) > 2 else "") or "",
            "sales": int(round(vendite or 0)),
        })
    rows.sort(key=lambda x: (-x["sales"], x["pdv"]))
    return {
        "file_name": path.name,
        "updated_at": updated_at,
        "rows": rows
    }


def pct(cur, prev):
    if cur is None or prev in (None,0):
        return None
    return (cur-prev)/prev


def load_lista(path):
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws=wb[wb.sheetnames[0]]
    mp={}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pdv=norm_pdv(row[0] if len(row)>0 else None)
        if not pdv: continue
        mp[pdv]={
            "agente": (row[9] if len(row)>9 else None) or "",
            "lista_area": row[1] if len(row)>1 else "",
            "lista_regione": row[4] if len(row)>4 else "",
            "lista_provincia": (row[5] if len(row)>5 else None) or (row[2] if len(row)>2 else "") or "",
            "lista_citta": (row[6] if len(row)>6 else None) or (row[3] if len(row)>3 else "") or "",
            "lista_indirizzo": (row[7] if len(row)>7 else None) or "",
        }
    return mp


def load_anag(path):
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws=wb[wb.sheetnames[0]]
    mp={}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pdv=norm_pdv(row[0] if len(row)>0 else None)
        if not pdv: continue
        mp[pdv]={
            "rzv": row[9] if len(row)>9 else "",
            "cr": row[10] if len(row)>10 else "",
            "mail_cr": row[11] if len(row)>11 else "",
            "filiale_commerciale": row[8] if len(row)>8 else "",
        }
    return mp


def find_report_sheet(wb):
    for s in wb.sheetnames:
        ws=wb[s]
        texts=[]
        for row in ws.iter_rows(min_row=1,max_row=6,min_col=1,max_col=25,values_only=True):
            texts.extend([str(v) for v in row if v not in (None,"")])
        j=" ".join(texts).upper()
        if "TELEPASS POINT BY ENI - REPORT ATTIVITA" in j:
            return ws
    return wb[wb.sheetnames[0]]


def extract_week_year(path):
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws=find_report_sheet(wb)
    texts=[]
    for row in ws.iter_rows(min_row=1,max_row=6,min_col=1,max_col=30,values_only=True):
        texts.extend([str(v) for v in row if v not in (None,"")])
    pats=[r"W\s*(\d{4})\s*/\s*(\d{1,2})",r"W\s*(\d{1,2})\s*/\s*(\d{4})",r"(20\d{2})\s*/\s*(\d{1,2})"]
    for t in texts:
        s=str(t)
        for pat in pats:
            m=re.search(pat,s,re.I)
            if m:
                a,b=int(m.group(1)),int(m.group(2))
                if a>=2000 and 1<=b<=53:
                    return a,b,ws.title
                if b>=2000 and 1<=a<=53:
                    return b,a,ws.title
    # fallback header row 4
    vals=[str(ws.cell(4,c).value or "") for c in range(1,60)]
    for s in vals:
        for m in re.finditer(r"(20\d{2})\s*/\s*(\d{1,2})",s):
            y=int(m.group(1)); w=int(m.group(2))
            if 1<=w<=53:
                return y,w,ws.title
    return None,None,ws.title


def parse_report_dynamic(path):
    year, week, sheet = extract_week_year(path)
    wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws=wb[sheet]
    headers=[ws.cell(4,c).value for c in range(1,60)]
    norm_headers=[(str(h).strip().upper() if h is not None else "") for h in headers]
    def find_exact(txt, after=0):
        txt=txt.upper()
        for i,h in enumerate(norm_headers, start=1):
            if i<=after: continue
            if h==txt:
                return i
        return None
    def find_contains(txt, after=0):
        txt=txt.upper()
        for i,h in enumerate(norm_headers, start=1):
            if i<=after: continue
            if txt in h:
                return i
        return None

    c_pdv=find_contains("PV ENI")
    c_area=find_exact("AREA COMM.")
    c_reg=find_exact("REGIONE")
    c_prov=find_exact("PROVINCIA")
    c_city=find_exact("CITTÀ") or find_exact("CITTA")
    c_addr=find_exact("INDIRIZZO")
    c_attivo=find_contains("ATTIVO")
    c_data=find_contains("DATA ATTIVAZIONE")
    c_vend_week=find_exact(f"VENDITE {year}/{week:02d}")
    c_vend_ly=find_exact(f"VENDITE {year-1}/{week:02d}")
    # first twin/business after vend ly
    c_twin=find_exact("DI CUI TWIN", after=(c_vend_ly or 0))
    c_bus_week=find_exact("DI CUI BUSINESS", after=(c_twin or c_vend_ly or 0))
    c_ass_week=find_exact(f"ASS. STRAD. EU VENDITE {year}/{week:02d}")
    c_ass_ly=find_exact(f"ASS. STRAD. EU VENDITE {year-1}/{week:02d}")
    c_sost_week=find_exact("SOST.", after=(c_ass_ly or c_ass_week or 0))
    c_up_eu_week=find_exact("UPSELL. EU", after=(c_sost_week or 0))
    c_sost_family_week=find_exact("SOST. FAMILY", after=(c_up_eu_week or c_sost_week or 0))
    c_tot_sales=find_exact(f"TOTALE VENDITE TELEPASS {year}")
    c_tot_sales_prev=find_exact(f"TOTALE VENDITE TELEPASS {year-1}")
    c_tot_twin=find_exact("TOTALE TWIN", after=(c_tot_sales_prev or c_tot_sales or 0))
    c_tot_bus=find_exact("DI CUI BUSINESS", after=(c_tot_twin or c_tot_sales_prev or 0))
    c_tot_ass=find_exact(f"TOTALE ASS. STRAD. {year}")
    c_tot_ass_prev=find_exact(f"TOTALE ASS. STRAD. {year-1}")
    c_tot_sost=find_exact("SOST.", after=(c_tot_ass_prev or c_tot_ass or 0))
    c_tot_up_eu=find_exact("UPSELL. EU", after=(c_tot_sost or 0))
    c_tot_sost_family=find_exact("SOST. FAMILY", after=(c_tot_up_eu or c_tot_sost or 0))
    # debug if missing
    required=[c_pdv,c_city,c_vend_week,c_vend_ly,c_twin,c_bus_week,c_ass_week,c_ass_ly,c_sost_week,c_up_eu_week,c_sost_family_week,c_tot_sales,c_tot_sales_prev,c_tot_twin,c_tot_bus,c_tot_ass,c_tot_ass_prev,c_tot_sost,c_tot_up_eu,c_tot_sost_family]
    if any(v is None for v in required):
        # print(path, required, headers[:55])
        pass
    recs=[]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row: continue
        pdv=norm_pdv(row[c_pdv-1] if c_pdv else None)
        if not pdv: continue
        vend_week=safe_num(row[c_vend_week-1]) if c_vend_week else None
        bus_week=safe_num(row[c_bus_week-1]) if c_bus_week else 0
        # fix strange negatives/fractions
        if bus_week is not None and (bus_week < 0 or abs(bus_week-round(bus_week))>1e-6 and abs(bus_week)<1):
            bus_week=0
        twin_week=safe_num(row[c_twin-1]) if c_twin else None
        if twin_week is not None and twin_week < 0: twin_week=0
        recs.append({
            "pdv": pdv,
            "week_year": year,
            "week_num": week,
            "period": f"{year}-W{week:02d}",
            "area_report": row[c_area-1] if c_area else "",
            "regione": row[c_reg-1] if c_reg else "",
            "provincia": row[c_prov-1] if c_prov else "",
            "citta": row[c_city-1] if c_city else "",
            "indirizzo": row[c_addr-1] if c_addr else "",
            "data_attivazione": row[c_data-1].strftime("%Y-%m-%d") if c_data and hasattr(row[c_data-1], "strftime") else (str(row[c_data-1]) if c_data and row[c_data-1] else ""),
            "attivo": row[c_attivo-1] if c_attivo else "",
            "vendite_settimana": vend_week or 0,
            "vendite_anno_prec_stessa_sett": safe_num(row[c_vend_ly-1]) if c_vend_ly else None,
            "twin_settimana": twin_week or 0,
            "business_vendite_settimana": bus_week or 0,
            "prospect_settimana": max((vend_week or 0) - (bus_week or 0), 0),
            "ass_settimana": safe_num(row[c_ass_week-1]) if c_ass_week else 0,
            "ass_anno_prec_stessa_sett": safe_num(row[c_ass_ly-1]) if c_ass_ly else None,
            "sost_settimana": safe_num(row[c_sost_week-1]) if c_sost_week else 0,
            "upgrade_eu_settimana": safe_num(row[c_up_eu_week-1]) if c_up_eu_week else 0,
            "sost_family_settimana": safe_num(row[c_sost_family_week-1]) if c_sost_family_week else 0,
            "tot_vendite_anno": safe_num(row[c_tot_sales-1]) if c_tot_sales else 0,
            "tot_vendite_anno_prec": safe_num(row[c_tot_sales_prev-1]) if c_tot_sales_prev else 0,
            "tot_twin_report": safe_num(row[c_tot_twin-1]) if c_tot_twin else 0,
            "tot_business_vendite_anno": safe_num(row[c_tot_bus-1]) if c_tot_bus else 0,
            "tot_ass_anno": safe_num(row[c_tot_ass-1]) if c_tot_ass else 0,
            "tot_ass_anno_prec": safe_num(row[c_tot_ass_prev-1]) if c_tot_ass_prev else 0,
            "tot_sost_anno": safe_num(row[c_tot_sost-1]) if c_tot_sost else 0,
            "tot_upgrade_eu_anno": safe_num(row[c_tot_up_eu-1]) if c_tot_up_eu else 0,
            "tot_sost_family_anno": safe_num(row[c_tot_sost_family-1]) if c_tot_sost_family else 0,
            "source_file": os.path.basename(path)
        })
    return recs


def style_sheet(ws):
    thin = Side(style="thin", color=LINE)
    for row in ws.iter_rows():
        for c in row:
            c.border = Border(bottom=thin)
            c.alignment = Alignment(vertical="center")
    ws.sheet_view.showGridLines = False


def add_table(ws, start_row, start_col, headers, data, table_name):
    for j, h in enumerate(headers, start_col):
        c = ws.cell(start_row, j, h)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(data, start_row + 1):
        for j, val in enumerate(row, start_col):
            ws.cell(i, j, val)
    if data:
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col+len(headers)-1)}{start_row+len(data)}"
        tab = Table(displayName=table_name[:25].replace(" ",""), ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)


def autosize(ws, max_width=28):
    for col in range(1, ws.max_column + 1):
        values=[]
        for row in range(1, min(ws.max_row, 300)+1):
            val=ws.cell(row,col).value
            if val is None: continue
            values.append(str(val))
        width=min(max(max((len(v) for v in values), default=8)+2, 10), max_width)
        ws.column_dimensions[get_column_letter(col)].width = width


def create_mobile_workbook(rows, out_path, title, current_week, filter_text):
    wb=Workbook()
    ws=wb.active
    ws.title='HOME'
    ws.sheet_view.showGridLines=False
    ws.merge_cells('A1:F1')
    ws['A1']=title
    ws['A1'].fill=PatternFill('solid', fgColor=BLUE)
    ws['A1'].font=Font(size=18,bold=True,color='FFFFFF')
    ws['A1'].alignment=Alignment(horizontal='center')
    ws.row_dimensions[1].height=28
    ws['A3']='Filtro'; ws['B3']=filter_text
    ws['A4']=f'Settimana attuale'; ws['B4']=f'W{current_week:02d}'
    # summary cards
    cards=[
        ('PDV', len(rows), LIGHT),
        ('Vendite 2026', sum(r['tot_vendite_anno'] or 0 for r in rows), 'EAF2FF'),
        ('Prospect 2026', sum(r['prospect_ytd_calc'] or 0 for r in rows), 'E8F7EC'),
        ('Assistenze 2026', sum(r['tot_ass_anno'] or 0 for r in rows), 'FFF4E5'),
    ]
    start_col=1
    for idx,(label,val,fill) in enumerate(cards):
        c=start_col+idx
        ws.cell(6,c,label)
        ws.cell(7,c,val)
        ws.cell(6,c).font=Font(bold=True,color=BLUE)
        ws.cell(7,c).font=Font(bold=True,size=20,color=BLUE)
        ws.cell(6,c).fill=PatternFill('solid',fgColor=fill.replace('#',''))
        ws.cell(7,c).fill=PatternFill('solid',fgColor=fill.replace('#',''))
        ws.cell(6,c).alignment=Alignment(horizontal='center')
        ws.cell(7,c).alignment=Alignment(horizontal='center')
    headers=['PDV','Città','Agente','CR','Vend 2026','Vend 2025','Ass 2026','Ass 2025','Prospect','Twin','Business','Sost anno','UP EU','Stato']
    data=[]
    for r in rows:
        data.append([r['pdv'],r['citta'],r['agente'],r['cr'],r['tot_vendite_anno'],r['tot_vendite_anno_prec'],r['tot_ass_anno'],r['tot_ass_anno_prec'],r['prospect_ytd_calc'],r['twin_ytd_calc'],r['business_ytd_calc'],r['tot_sost_family_anno'],r['tot_upgrade_eu_anno'],r['stato']])
    add_table(ws,10,1,headers,data,'Report')
    for row in range(11, ws.max_row+1):
        for c in range(5,14):
            ws.cell(row,c).number_format='#,##0'
    ws.freeze_panes='A10'
    style_sheet(ws); autosize(ws, max_width=22)
    wb.save(out_path)


def safe_filename(s):
    s = re.sub(r"[\\/:*?\"<>|]+", " ", s or "").strip()
    s = re.sub(r"\s+", " ", s).strip()
    return s or "Senza_Nome"


def build_export_reports(base_dir, current_rows, current_week):
    export_dir = Path(base_dir)/'files'/'REPORT_FILTRI'
    export_dir.mkdir(parents=True, exist_ok=True)
    manifest={}
    groups=[('all','Tutti i punti vendita', current_rows, {"agente":"","cr":"","rzv":""})]
    by_agent=defaultdict(list); by_cr=defaultdict(list); by_rzv=defaultdict(list); by_pair=defaultdict(list)
    for r in current_rows:
        if r.get('agente'): by_agent[r['agente']].append(r)
        if r.get('cr'): by_cr[r['cr']].append(r)
        if r.get('rzv'): by_rzv[r['rzv']].append(r)
        if r.get('agente') or r.get('cr'): by_pair[(r.get('agente') or '', r.get('cr') or '')].append(r)
    for agent, rows in by_agent.items(): groups.append((f'agent::{agent}', f'Agente - {agent}', rows, {'agente':agent,'cr':'','rzv':''}))
    for cr, rows in by_cr.items(): groups.append((f'cr::{cr}', f'CR - {cr}', rows, {'agente':'','cr':cr,'rzv':''}))
    for rzv, rows in by_rzv.items(): groups.append((f'rzv::{rzv}', f'RZV - {rzv}', rows, {'agente':'','cr':'','rzv':rzv}))
    for (agent,cr), rows in by_pair.items(): groups.append((f'pair::{agent}||{cr}', f'Agente {agent} - CR {cr}', rows, {'agente':agent,'cr':cr,'rzv':''}))
    for key,title,rows,filt in groups:
        fname=safe_filename(title)+'.xlsx'
        out_path=export_dir/fname
        filt_text=' · '.join([f'Agente: {filt["agente"]}' if filt.get('agente') else '',
                              f'CR: {filt["cr"]}' if filt.get('cr') else '',
                              f'RZV: {filt["rzv"]}' if filt.get('rzv') else '']).strip(' ·')
        if not filt_text: filt_text='Tutti i punti vendita'
        create_mobile_workbook(rows, str(out_path), title, current_week, filt_text)
        manifest[key]={"path": f"files/REPORT_FILTRI/{fname}", "title": title, "rows": len(rows)}
    return manifest


def build_data_for_html(current_rows, hist, summary, export_manifest, file_utili, current_week, current_year, gara_live=None):
    data_rows=[]
    hist_map={}
    for r in current_rows:
        ops=(r.get('tot_vendite_anno') or 0)+(r.get('tot_sost_family_anno') or 0)
        # recent 4 weeks
        harr=hist[r['pdv']]
        recent=harr[-4:]
        recent_comp=[{
            "week": x['week_num'],
            "sales_total": int(round(x.get('vendite_settimana') or 0)),
            "prospect": int(round(x.get('prospect_settimana') or 0)),
            "business": int(round(x.get('business_vendite_settimana') or 0)),
            "assist": int(round(x.get('ass_settimana') or 0)),
            "twin": int(round(x.get('twin_settimana') or 0)),
            "sost_family": int(round(x.get('sost_family_settimana') or 0)),
            "up_eu": int(round(x.get('upgrade_eu_settimana') or 0)),
        } for x in recent]
        hist_map[r['pdv']]=[{
            "week": x['week_num'],
            "sales_total": int(round(x.get('vendite_settimana') or 0)),
            "prospect": int(round(x.get('prospect_settimana') or 0)),
            "twin": int(round(x.get('twin_settimana') or 0)),
            "business": int(round(x.get('business_vendite_settimana') or 0)),
            "assist": int(round(x.get('ass_settimana') or 0)),
            "sost_family": int(round(x.get('sost_family_settimana') or 0)),
            "up_eu": int(round(x.get('upgrade_eu_settimana') or 0)),
        } for x in harr]
        data_rows.append({
            "pdv": r['pdv'],
            "city": r.get('citta',''),
            "address": r.get('indirizzo',''),
            "agent": r.get('agente',''),
            "cr": r.get('cr',''),
            "rzv": r.get('rzv',''),
            "status": r.get('stato',''),
            "rank_sales": r.get('rank_all'),
            "rank_text": r.get('rank_text'),
            "latest_week": r.get('week_num'),
            "latest": {
                "sales_total": int(round(r.get('vendite_settimana') or 0)),
                "prospect": int(round(r.get('prospect_settimana') or 0)),
                "business": int(round(r.get('business_vendite_settimana') or 0)),
                "assist": int(round(r.get('ass_settimana') or 0)),
                "twin": int(round(r.get('twin_settimana') or 0)),
                "sost_family": int(round(r.get('sost_family_settimana') or 0)),
                "up_eu": int(round(r.get('upgrade_eu_settimana') or 0))
            },
            "prev": {
                "week": r.get('prev_week'),
                "sales_total": int(round((r.get('vendite_settimana') or 0) - (r.get('vendite_week_diff') or 0))) if r.get('prev_week') else None,
                "prospect": int(round((r.get('prospect_settimana') or 0) - (r.get('prospect_week_diff') or 0))) if r.get('prev_week') else None,
                "assist": int(round((r.get('ass_settimana') or 0) - (r.get('ass_week_diff') or 0))) if r.get('prev_week') else None,
            },
            "ytd": {
                "sales_2026": int(round(r.get('tot_vendite_anno') or 0)),
                "sales_2025": int(round(r.get('tot_vendite_anno_prec') or 0)),
                "sales_pct": round((r.get('sales_ytd_pct') or 0)*100,1) if r.get('sales_ytd_pct') is not None else None,
                "prospect": int(round(r.get('prospect_ytd_calc') or 0)),
                "twin": int(round(r.get('twin_ytd_calc') or 0)),
                "business": int(round(r.get('business_ytd_calc') or 0)),
                "assist_2026": int(round(r.get('tot_ass_anno') or 0)),
                "assist_2025": int(round(r.get('tot_ass_anno_prec') or 0)),
                "assist_pct": round((r.get('assist_ytd_pct') or 0)*100,1) if r.get('assist_ytd_pct') is not None else None,
                "sost_family": int(round(r.get('tot_sost_family_anno') or 0)),
                "up_eu": int(round(r.get('tot_upgrade_eu_anno') or 0)),
                "attach_rate": round((r.get('attach_rate') or 0)*100,1) if r.get('attach_rate') is not None else None,
                "up_eu_rate": round((r.get('up_eu_rate') or 0)*100,1) if r.get('up_eu_rate') is not None else None,
                "recovery_week_need": round(r.get('sales_recovery_weekly_need') or 0,1),
                "current_weekly_avg": round(r.get('current_weekly_avg') or 0,1),
            },
            "flags": {
                "follow": r.get('stato')!='Bene',
                "sales_below": ((r.get('sales_ytd_pct') or 0)<=-0.05 and (r.get('sales_ytd_diff') or 0)<=-10),
                "assist_below": ((r.get('assist_ytd_pct') or 0)<=-0.10 and (r.get('assist_ytd_diff') or 0)<=-5),
                "recent_decline": bool(r.get('trend_note'))
            },
            "notes": r.get('motivi') or ['Andamento regolare'],
            "recent": recent_comp
        })
    data={
        "meta":{"current_week": current_week, "current_year": current_year, "generated_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M")},
        "summary": summary,
        "rows": data_rows,
        "history": hist_map,
        "export_manifest": export_manifest,
        "file_utili": file_utili,
        "gara_live": gara_live,
        "gare_pdv":[],
        "gare_agenti":[]
    }
    return data



def load_config():
    default = {
        "year_mode": "latest_year_only",
        "thresholds": {
            "sales_bad_pct": -0.15,
            "sales_warn_pct": -0.05,
            "sales_bad_abs": 30,
            "sales_warn_abs": 10,
            "assist_bad_pct": -0.15,
            "assist_warn_pct": -0.10,
            "assist_bad_abs": 10,
            "assist_warn_abs": 5
        },
        "gare_pdv": [],
        "gare_agenti": []
    }
    if not os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(default, f, ensure_ascii=False, indent=2)
        return default
    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as f:
            user = json.load(f)
        if not isinstance(user, dict):
            return default
        for k, v in default.items():
            if k not in user:
                user[k] = v
        return user
    except Exception:
        return default

def show_popup(title, text):
    if tk is None:
        return
    try:
        messagebox.showinfo(title, text)
    except Exception:
        pass

def pick_inputs():
    if tk is None:
        print("Tkinter non disponibile. Avvia con: python aggiorna_dashboard.py <lista_pdv.xlsx> <anagrafica.xlsx> <cartella_report> <cartella_output>")
        sys.exit(1)

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    show_popup(
        "Dashboard Telepass ENI",
        "Ti guiderò passo passo.\n\n"
        "1) Seleziona LISTA PDV ENI\n"
        "2) Seleziona ANAGRAFICA aggiornata\n"
        "3) Seleziona la cartella con i report settimana\n"
        "4) Seleziona la cartella OUTPUT\n\n"
        "Uso solo l'anno più recente trovato dentro i report."
    )

    show_popup("Passo 1 di 4", "Ora scegli il file LISTA PDV ENI.")
    lista = filedialog.askopenfilename(title="Passo 1 di 4 - LISTA PDV ENI", filetypes=[("Excel", "*.xlsx")])
    if not lista: sys.exit()

    show_popup("Passo 2 di 4", "Ora scegli il file ANAGRAFICA aggiornata.")
    anag = filedialog.askopenfilename(title="Passo 2 di 4 - ANAGRAFICA aggiornata", filetypes=[("Excel", "*.xlsx")])
    if not anag: sys.exit()

    show_popup("Passo 3 di 4", "Ora scegli la CARTELLA con tutte le settimane.")
    report_dir = filedialog.askdirectory(title="Passo 3 di 4 - CARTELLA report settimane")
    if not report_dir: sys.exit()

    show_popup("Passo 4 di 4", "Ora scegli la cartella OUTPUT dove vuoi salvare il risultato.")
    out_dir = filedialog.askdirectory(title="Passo 4 di 4 - CARTELLA OUTPUT")
    if not out_dir: sys.exit()

    return lista, anag, report_dir, out_dir

def scan_report_files(report_dir, year_mode="latest_year_only"):
    found = []
    skipped = []
    for root, _, files in os.walk(report_dir):
        for name in files:
            if not name.lower().endswith(".xlsx") or name.startswith("~$"):
                continue
            path = os.path.join(root, name)
            try:
                year, week, _ = extract_week_year(path)
                if year and week:
                    found.append((path, year, week, os.path.getmtime(path)))
                else:
                    skipped.append((path, "Settimana non trovata"))
            except Exception as e:
                skipped.append((path, f"Errore lettura: {e}"))
    if not found:
        return {"selected_paths": [], "selected_keys": [], "skipped": skipped, "selected_year": None, "missing_weeks": []}
    years = sorted(set(y for _, y, _, _ in found))
    selected_year = max(years) if year_mode == "latest_year_only" else None
    if selected_year is not None:
        found = [x for x in found if x[1] == selected_year]
    best = {}
    for path, year, week, mtime in found:
        key = (year, week)
        if key not in best or mtime > best[key][1]:
            best[key] = (path, mtime)
    keys = sorted(best)
    weeks = [w for y, w in keys]
    missing = []
    if weeks:
        for w in range(min(weeks), max(weeks)+1):
            if w not in weeks:
                missing.append((selected_year, w))
    return {"selected_paths": [best[k][0] for k in keys], "selected_keys": keys, "skipped": skipped, "selected_year": selected_year, "missing_weeks": missing}

def enrich_current(records, config):
    hist = defaultdict(list)
    for r in records:
        hist[r["pdv"]].append(r)
    for pdv in hist:
        hist[pdv].sort(key=lambda x: (x["week_year"], x["week_num"]))
    current_yearweek = max((r["week_year"], r["week_num"]) for r in records)
    current = []
    th = config["thresholds"]
    for pdv, arr in hist.items():
        cur = next((x for x in arr if (x["week_year"], x["week_num"]) == current_yearweek), None)
        if not cur:
            continue
        cur = cur.copy()
        prev = arr[-2] if len(arr) >= 2 else None
        cur["agente"] = cur.get("agente", "")
        cur["cr"] = cur.get("cr", "")
        cur["rzv"] = cur.get("rzv", "")
        cur["business_ytd_calc"] = sum((x.get("business_vendite_settimana") or 0) for x in arr)
        cur["twin_ytd_calc"] = sum((x.get("twin_settimana") or 0) for x in arr)
        cur["prospect_ytd_calc"] = sum((x.get("prospect_settimana") or 0) for x in arr)
        cur["prev_week"] = prev["week_num"] if prev else None
        cur["vendite_week_diff"] = (cur.get("vendite_settimana") or 0) - ((prev or {}).get("vendite_settimana") or 0) if prev else None
        cur["prospect_week_diff"] = (cur.get("prospect_settimana") or 0) - ((prev or {}).get("prospect_settimana") or 0) if prev else None
        cur["ass_week_diff"] = (cur.get("ass_settimana") or 0) - ((prev or {}).get("ass_settimana") or 0) if prev else None
        cur["sales_ytd_diff"] = (cur.get("tot_vendite_anno") or 0) - (cur.get("tot_vendite_anno_prec") or 0)
        cur["assist_ytd_diff"] = (cur.get("tot_ass_anno") or 0) - (cur.get("tot_ass_anno_prec") or 0)
        cur["sales_ytd_pct"] = pct(cur.get("tot_vendite_anno"), cur.get("tot_vendite_anno_prec"))
        cur["assist_ytd_pct"] = pct(cur.get("tot_ass_anno"), cur.get("tot_ass_anno_prec"))
        ops = (cur.get("tot_vendite_anno") or 0) + (cur.get("tot_sost_family_anno") or 0)
        cur["attach_rate"] = (cur.get("tot_ass_anno") or 0) / ops if ops else None
        cur["up_eu_rate"] = (cur.get("tot_upgrade_eu_anno") or 0) / (cur.get("tot_sost_family_anno") or 0) if (cur.get("tot_sost_family_anno") or 0) else None
        remaining = max(52 - cur["week_num"], 1)
        gap = max((cur.get("tot_vendite_anno_prec") or 0) - (cur.get("tot_vendite_anno") or 0), 0)
        cur["sales_recovery_weekly_need"] = gap / remaining if gap > 0 else 0
        cur["current_weekly_avg"] = (cur.get("tot_vendite_anno") or 0) / max(cur["week_num"], 1)

        trend = ""
        if len(arr) >= 2 and (arr[-1].get("vendite_settimana") or 0) < (arr[-2].get("vendite_settimana") or 0):
            if len(arr) >= 3 and (arr[-2].get("vendite_settimana") or 0) < (arr[-3].get("vendite_settimana") or 0):
                trend = f"In calo da 2 settimane (W{arr[-3]['week_num']:02d} → W{arr[-2]['week_num']:02d} → W{arr[-1]['week_num']:02d})"
            else:
                trend = f"Ultima settimana in calo vs W{arr[-2]['week_num']:02d}"
        cur["trend_note"] = trend

        sp = cur["sales_ytd_pct"] if cur["sales_ytd_pct"] is not None else 0
        ap = cur["assist_ytd_pct"] if cur["assist_ytd_pct"] is not None else 0
        sales_bad = sp <= th["sales_bad_pct"] and cur["sales_ytd_diff"] <= -th["sales_bad_abs"]
        sales_warn = sp <= th["sales_warn_pct"] and cur["sales_ytd_diff"] <= -th["sales_warn_abs"]
        assist_bad = ap <= th["assist_bad_pct"] and cur["assist_ytd_diff"] <= -th["assist_bad_abs"]
        assist_warn = ap <= th["assist_warn_pct"] and cur["assist_ytd_diff"] <= -th["assist_warn_abs"]

        reasons = []
        if sales_bad or sales_warn:
            reasons.append("Vendite 2026 sotto il 2025")
        if assist_bad or assist_warn:
            reasons.append("Assistenze 2026 sotto il 2025")
        if trend:
            reasons.append(trend)
        if not reasons:
            reasons.append("Andamento regolare")

        if sales_bad or (sales_warn and assist_warn):
            stato = "Male"
        elif sales_warn or assist_warn or trend:
            stato = "Da seguire"
        else:
            stato = "Bene"
        cur["stato"] = stato
        cur["motivi"] = reasons
        current.append(cur)

    current.sort(key=lambda r: ((r.get("tot_vendite_anno") or 0), (r.get("prospect_ytd_calc") or 0)), reverse=True)
    total = len(current)
    for i, r in enumerate(current, start=1):
        r["rank_all"] = i
        r["rank_text"] = f"{i} su {total}"
    return current_yearweek, current, hist

def build_summary(current):
    return {
        "total_sales_2026": int(round(sum(r.get("tot_vendite_anno") or 0 for r in current))),
        "total_sales_2025": int(round(sum(r.get("tot_vendite_anno_prec") or 0 for r in current))),
        "total_prospect_2026": int(round(sum(r.get("prospect_ytd_calc") or 0 for r in current))),
        "total_ass_2026": int(round(sum(r.get("tot_ass_anno") or 0 for r in current))),
        "total_ass_2025": int(round(sum(r.get("tot_ass_anno_prec") or 0 for r in current))),
        "pdv_count": len(current),
    }

def copy_file_utili(out_dir):
    dest = os.path.join(out_dir, "files")
    os.makedirs(dest, exist_ok=True)
    files = []
    if os.path.isdir(FILE_UTILI_DIR):
        for name in os.listdir(FILE_UTILI_DIR):
            src = os.path.join(FILE_UTILI_DIR, name)
            if os.path.isfile(src):
                out = os.path.join(dest, name)
                shutil.copy2(src, out)
                files.append({"name": name, "path": f"files/{name}"})
    return files

def build_html(data):
    with open(TEMPLATE_PATH, "r", encoding="utf-8") as f:
        tpl = f.read()
    return tpl.replace("__DATA_JSON__", json.dumps(data, ensure_ascii=False)).replace("__CURRENT_WEEK__", f"{data['meta']['current_week']:02d}")

def build_master_workbook(out_path, current, records, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "LEGGIMI"
    ws["A1"] = "Dashboard Telepass ENI"
    ws["A1"].font = Font(size=16, bold=True, color=BLUE)
    ws["A3"] = "Apri il file HTML per la dashboard."
    ws["A4"] = "Questo Excel è l'archivio dati."
    ws["A6"] = f"Anno usato: {meta['selected_year']}"
    ws["A7"] = f"Settimana attuale: W{meta['current_week']:02d}"
    ws["A8"] = f"PDV attuali: {len(current)}"
    ws["A9"] = f"Righe storico: {len(records)}"
    ws.column_dimensions["A"].width = 70

    headers = ["pdv","citta","indirizzo","agente","cr","rzv","stato","rank_all","vendite_settimana","prospect_settimana","business_vendite_settimana","twin_settimana","tot_vendite_anno","tot_vendite_anno_prec","tot_ass_anno","tot_ass_anno_prec","twin_ytd_calc","business_ytd_calc","prospect_ytd_calc","tot_sost_family_anno","tot_upgrade_eu_anno","motivi"]
    ws2 = wb.create_sheet("PDV_Attuali")
    data = []
    for r in current:
        data.append([
            r.get("pdv"), r.get("citta"), r.get("indirizzo"), r.get("agente"), r.get("cr"), r.get("rzv"), r.get("stato"), r.get("rank_all"),
            r.get("vendite_settimana"), r.get("prospect_settimana"), r.get("business_vendite_settimana"), r.get("twin_settimana"),
            r.get("tot_vendite_anno"), r.get("tot_vendite_anno_prec"), r.get("tot_ass_anno"), r.get("tot_ass_anno_prec"),
            r.get("twin_ytd_calc"), r.get("business_ytd_calc"), r.get("prospect_ytd_calc"), r.get("tot_sost_family_anno"), r.get("tot_upgrade_eu_anno"),
            " | ".join(r.get("motivi") or [])
        ])
    add_table(ws2, 1, 1, headers, data, "PDVAttuali")
    ws3 = wb.create_sheet("Storico")
    hist_headers = ["pdv","period","week_num","vendite_settimana","prospect_settimana","business_vendite_settimana","twin_settimana","ass_settimana","sost_family_settimana","upgrade_eu_settimana","tot_vendite_anno","tot_ass_anno","source_file"]
    hist_rows = [[r.get(h) for h in hist_headers] for r in records]
    add_table(ws3, 1, 1, hist_headers, hist_rows, "Storico")
    for wsx in [ws, ws2, ws3]:
        style_sheet(wsx)
        autosize(wsx, max_width=34)
    wb.save(out_path)

def main():
    config = load_config()
    if len(sys.argv) >= 5:
        lista, anag, report_dir, out_dir = sys.argv[1:5]
    else:
        lista, anag, report_dir, out_dir = pick_inputs()

    os.makedirs(out_dir, exist_ok=True)
    lista_map = load_lista(lista)
    anag_map = load_anag(anag)

    scan = scan_report_files(report_dir, year_mode=config.get("year_mode", "latest_year_only"))
    if not scan["selected_paths"]:
        raise RuntimeError("Nessun report ENI valido trovato nella cartella selezionata.")

    records = []
    for path in scan["selected_paths"]:
        for r in parse_report_dynamic(path):
            li = lista_map.get(r["pdv"], {})
            an = anag_map.get(r["pdv"], {})
            r["agente"] = li.get("agente", "") or ""
            r["rzv"] = an.get("rzv", "") or ""
            r["cr"] = an.get("cr", "") or ""
            if not r["citta"]:
                r["citta"] = li.get("lista_citta", "")
            if not r["indirizzo"]:
                r["indirizzo"] = li.get("lista_indirizzo", "")
            records.append(r)
    ded = {}
    for r in records:
        ded[(r["pdv"], r["week_year"], r["week_num"])] = r
    records = sorted(ded.values(), key=lambda x: (x["week_year"], x["week_num"], x["pdv"]))

    current_yearweek, current, hist = enrich_current(records, config)
    summary = build_summary(current)

    os.makedirs(os.path.join(out_dir, "files"), exist_ok=True)
    export_manifest = build_export_reports(out_dir, current, current_yearweek[1])
    file_utili = copy_file_utili(out_dir)
    gara_live = load_live_gara()

    data = build_data_for_html(current, hist, summary, export_manifest, file_utili, current_yearweek[1], current_yearweek[0], gara_live)
    data["gare_pdv"] = list(config.get("gare_pdv", []))
    if gara_live and gara_live.get("rows"):
        live_rows = []
        row_map = {r["pdv"]: r for r in current}
        for i, gr in enumerate(gara_live["rows"], start=1):
            base = row_map.get(gr["pdv"], {})
            live_rows.append({
                "rank": i,
                "pdv": gr["pdv"],
                "city": base.get("citta") or gr.get("city",""),
                "agent": base.get("agente",""),
                "cr": base.get("cr",""),
                "rzv": base.get("rzv") or gr.get("rzv",""),
                "sales": gr["sales"]
            })
        data["gare_pdv"].insert(0, {
            "title": "Gara Family Base · andamento attuale",
            "updated_at": gara_live.get("updated_at",""),
            "file_name": gara_live.get("file_name",""),
            "items": [
                "Aggiornamento automatico dal file gara caricato nella cartella input/gare.",
                "Il numero indica a quante vendite è il PDV alla data del file."
            ],
            "table": live_rows
        })
    data["gare_agenti"] = config.get("gare_agenti", [])

    html_path = os.path.join(out_dir, "Telepass_ENI_sito_v6.html")
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(build_html(data))

    master_xlsx = os.path.join(out_dir, "Dati_Telepass_ENI_v6.xlsx")
    build_master_workbook(master_xlsx, current, records, {"selected_year": scan["selected_year"], "current_week": current_yearweek[1]})

    log_path = os.path.join(out_dir, "log_file_usati_v6.txt")
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("FILE REPORT TROVATI E USATI\n")
        f.write("==========================\n")
        for p in scan["selected_paths"]:
            y, w, _ = extract_week_year(p)
            f.write(f"{y}/W{w:02d} -> {p}\n")
        f.write("\nSETTIMANE MANCANTI\n")
        f.write("==================\n")
        if scan["missing_weeks"]:
            for y, w in scan["missing_weeks"]:
                f.write(f"{y}/W{w:02d}\n")
        else:
            f.write("Nessuna settimana mancante nel blocco usato.\n")
        f.write("\nFILE SCARTATI\n")
        f.write("============\n")
        for p, reason in scan["skipped"]:
            f.write(f"{p} -> {reason}\n")

    print("Creato:", html_path)
    print("Creato:", master_xlsx)
    print("Creato:", log_path)

if __name__ == "__main__":
    main()
